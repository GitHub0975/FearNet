#!/usr/bin/env python
# coding: utf-8

# # DEMO ON MNIST SPLIT


from MAS import *
from CIFAR_split import *
import shutil
import matplotlib.ticker as ticker
import torch.nn as nn

data_path = 'D:\\Brain_Scan_PCA'
batch_size=100
kwargs = {'num_workers': 0, 'pin_memory': True} 


def init_weights(m):
    if type(m) == nn.Linear:
        nn.init.xavier_normal(m.weight)
        m.bias.data.fill_(0.01)



# 將MNIST dataset切成五個任務的資料儲存(放進loader就可以用了)
# 建立TASK1~TASK50
class_split = [[i for i in range(50)]]
# 5類算一個task
for i in range(50, 100, 5):
    #print(np.arange(i,i+10))
    class_split.append([j for j in range(i, i+5)])
print(class_split)
#exit()
'''
task = 1
for digits in class_split:
    #print(digits)                
    # 使用label的真值(0-9作為label)
    dsets = {}
    dsets['train_true']=    CIFAR_Split(data_path, train=True, download=True,
                       transform=transforms.Compose([
                           transforms.ToTensor(),
                           transforms.Normalize(mean = [0.485, 0.456, 0.406], std = [0.229, 0.224, 0.225])
                           #transforms.Normalize(mean=[0.507, 0.487, 0.441], std=[0.267, 0.256, 0.276])
                       ]),digits=digits)
    
    dsets['val_true']=  CIFAR_Split(data_path, train=False, transform=transforms.Compose([
                           transforms.ToTensor(),
                           transforms.Normalize(mean = [0.485, 0.456, 0.406], std = [0.229, 0.224, 0.225])
                           #transforms.Normalize(mean=[0.507, 0.487, 0.441], std=[0.267, 0.256, 0.276])
                       ]),digits=digits)
                
    #print(dsets['train_true'].classes)
          
    dlabel=str(task)
    
    if not os.path.isdir('data/Pytorch_ImageNet_dataset_normalize'):
        os.mkdir('data/Pytorch_ImageNet_dataset_normalize')
    torch.save(dsets,'data/Pytorch_ImageNet_dataset_normalize/split'+dlabel+'_dataset.pth.tar')
    task += 1
exit()
'''
task = 1

#FIRST TASK TRAINING
from MNIST_Net import *
test = Fearnet()
#test.apply(init_weights)
torch.save(test, 'General_utils/Fearnet.pth.tar')

model_path='General_utils/Fearnet.pth.tar'
from Finetune_SGD import *
#digits = [1,2]
digits = class_split[0]
dlabel=str(task)


dataset_path=data_path + '/data/Pytorch_ImageNet_dataset_normalize/split'+dlabel+'_dataset.pth.tar'

exp_dir='exp_dir/Two_phase/cifar_NET_normalize'+dlabel

num_epochs=30

fine_tune_SGD(dataset_path=dataset_path, num_epochs=num_epochs,exp_dir=exp_dir,model_path=model_path,lr=1e-4,batch_size=64, class_num=len(digits))
#model_path=os.path.join(exp_dir,'best_model.pth.tar')

# 把50個類別的dataset檔名加到list中
'''task = 1
for digits in class_split[1:]:
    dlabel = str(task)
    dataset_path='data/Pytorch_ImageNet_dataset_normalize//split'+dlabel+'_dataset.pth.tar'
    reg_sets.append(dataset_path)
    task += 1
    #print(dataset_path)'''
    

#MIMIC the case when samples from the previous takss are seen in each step
from MAS import *
'''
if os.path.isfile('./General_utils/current_feature_mean_copy.npy'):

    os.remove('./General_utils/current_feature_mean.npy')
    os.remove('./General_utils/current_feature_cov.npy')
    os.remove('./General_utils/prev_feature_mean.npy')
    os.remove('./General_utils/prev_feature_cov.npy')
    
    shutil.copyfile('./General_utils/current_feature_mean_copy.npy', './General_utils/current_feature_mean.npy')
    shutil.copyfile('./General_utils/current_feature_cov_copy.npy', './General_utils/current_feature_cov.npy')
    shutil.copyfile('./General_utils/current_feature_cov_copy.npy', './General_utils/prev_feature_mean.npy')
    shutil.copyfile('./General_utils/current_feature_cov_copy.npy', './General_utils/prev_feature_cov.npy')
    
else:
    shutil.copyfile('./General_utils/current_feature_mean.npy', './General_utils/current_feature_mean_copy.npy')
    shutil.copyfile('./General_utils/current_feature_cov.npy', './General_utils/current_feature_cov_copy.npy')
    shutil.copyfile('./General_utils/prev_feature_mean.npy', './General_utils/prev_feature_mean_copy.npy')
    shutil.copyfile('./General_utils/prev_feature_cov.npy', './General_utils/prev_feature_cov_copy.npy')
'''
reg_sets=[]
dataset_path = 'data/Pytorch_ImageNet_dataset_normalize//split1_dataset.pth.tar'

exp_dir='exp_dir/Two_phase/cifar_NET_normalize1'
#exp_dir='exp_dir/Two_phase/cifar_NET_normalize7'
previous_pathes=[]
reg_lambda=10

history_task_acc=np.zeros((51, 51))
history_task_acc[0][0] = 0.71    # 之後要再改寫
#task = 7
for digits in class_split[1:]:
    reg_sets.append(dataset_path)    # 舊資料

    model_path=os.path.join(exp_dir,'epoch.pth.tar')
    previous_pathes.append(model_path)    # 每次的新任務都加上上一次訓練好的模型
    print(model_path)
    
    task += 1
    #task = 50
    dlabel=str(task)

    dataset_path = 'data/Pytorch_ImageNet_dataset_normalize//split'+dlabel+'_dataset.pth.tar'    # 新任務的資料路徑

    exp_dir='exp_dir/Two_phase/cifar_NET_normalize'+dlabel
   
    num_epochs=10
    
    data_dirs=None
    
    dsets = torch.load(dataset_path)

    class_num = digits[-1] + 1     # 目前有幾個類別

    print('Total training class number: {}'.format(class_num))
    
    if task % 2 == 0:
        
        if task != 2:
            print('----------Big_Merge implement--------------------')
            model_ft = Big_Merger(dataset_path=dataset_path, previous_pathes=previous_pathes, 
                previous_task_model_path=model_path, exp_dir=exp_dir, data_dirs=data_dirs, 
                reg_sets=reg_sets, reg_lambda=reg_lambda, batch_size=64, num_epochs=10, lr=1e-4,norm='L2', 
                b1=False, class_num = class_num, task = task, digits = digits)
        print('----------Big_Separate implement--------------------')        
        if os.path.isfile(os.path.join(exp_dir, 'Merge.pth.tar')):
            model_path=os.path.join(exp_dir,'Merge.pth.tar')
            print(model_path)
        model_ft, task_acc = Big_Separate(dataset_path=dataset_path, previous_pathes=previous_pathes, 
                previous_task_model_path=model_path, exp_dir=exp_dir, data_dirs=data_dirs, 
                reg_sets=reg_sets, reg_lambda=reg_lambda, batch_size=64, num_epochs=num_epochs, lr=1e-4,norm='L2', 
                b1=False, class_num = class_num, task = task, digits = digits)
        #exit()
    else:
        #num_epochs = 20
        print('----------Train New implement--------------------')
        model_ft, task_acc = Train_New(dataset_path=dataset_path, previous_pathes=previous_pathes, 
                previous_task_model_path=model_path, exp_dir=exp_dir, data_dirs=data_dirs, 
                reg_sets=reg_sets, reg_lambda=reg_lambda, batch_size=64, num_epochs=num_epochs, lr=1e-4,norm='L2', 
                b1=False, class_num = class_num, task = task, digits = digits)
    
    # 紀錄每個task的歷史紀錄
    for i in range(len(task_acc)):
        history_task_acc[i][task-1] = task_acc[i]
    
    print("Final Model's state_dict:")
    #print(model_ft)
    layer_names = []
    #for param_tensor in model_ft.state_dict():
    #    layer_names.append(param_tensor)
    #    print(param_tensor, "\t", model_ft.state_dict()[param_tensor].size())
    
    
    
# 儲存歷史記錄圖
fig, ax = plt.subplots(2,2)
plt.xticks(fontsize=20)

fig.set_figheight(25)
fig.set_figwidth(70)

        
colormap = plt.cm.gist_ncar
colors = [colormap(i) for i in np.random.rand(12)]

for i in range(12):
    legend_label = 'task' + str(i)
    ax[0, 0].plot(history_task_acc[i], 'o--', color = colors[i], label = legend_label)
ax[0,0].legend(bbox_to_anchor=(0.5, 1.15), ncol=3)
ax[0,0].set_xlabel('number of tasks')
ax[0,0].axis(xmin=0, xmax=51)
ax[0,0].axis(ymin=0, ymax=1)

ax[0,0].xaxis.set_major_locator(ticker.MultipleLocator(1))
ax[0,0].yaxis.set_major_locator(ticker.MultipleLocator(0.1))
ax[0,0].grid(True)

colors = [colormap(i) for i in np.random.rand(12)]
for i in range(12, 24):
    legend_label = 'task' + str(i)
    ax[0, 1].plot(history_task_acc[i], 'o--', color = colors[i-12], label = legend_label)
ax[0,1].legend(bbox_to_anchor=(0.5, 1.15), ncol=3)
ax[0,1].set_xlabel('number of tasks')
ax[0,1].axis(xmin=0, xmax=51)
ax[0,1].axis(ymin=0, ymax=1)

ax[0,1].xaxis.set_major_locator(ticker.MultipleLocator(1))
ax[0,1].yaxis.set_major_locator(ticker.MultipleLocator(0.1))
ax[0,1].grid(True)

colors = [colormap(i) for i in np.random.rand(12)]
for i in range(24, 36):
    legend_label = 'task' + str(i)
    ax[1, 0].plot(history_task_acc[i], 'o--', color = colors[i-24], label = legend_label)
ax[1,0].legend(bbox_to_anchor=(0.5, 1.15), ncol=3)
ax[1,0].set_xlabel('number of tasks')
ax[1,0].axis(xmin=0, xmax=51)
ax[1,0].axis(ymin=0, ymax=1)

ax[1,0].xaxis.set_major_locator(ticker.MultipleLocator(1))
ax[1,0].yaxis.set_major_locator(ticker.MultipleLocator(0.1))
ax[1,0].grid(True)

colors = [colormap(i) for i in np.random.rand(15)]
for i in range(36, len(history_task_acc)):
    legend_label = 'task' + str(i)
    ax[1, 1].plot(history_task_acc[i], 'o--', color = colors[i-36], label = legend_label)
ax[1,1].legend(bbox_to_anchor=(0.5, 1.15), ncol=3)
ax[1,1].set_xlabel('number of tasks')
ax[1,1].axis(xmin=0, xmax=51)
ax[1,1].axis(ymin=0, ymax=1)

ax[1,1].xaxis.set_major_locator(ticker.MultipleLocator(1))
ax[1,1].yaxis.set_major_locator(ticker.MultipleLocator(0.1))
ax[1,1].grid(True)

#plt.savefig('./Distribution_Figure/task_acc_trend.png')
# In[24]:
# Avg_acc
Avg_acc = []
for i in range(task):
    Avg_acc.append(np.sum(np.array(history_task_acc)[:,i]) / (i + 1))
    #print(Avg_acc)
plt.figure(figsize=(70,25))
plt.plot(Avg_acc, 'o--')
plt.axis(xmin=0, xmax=51)
plt.grid()
plt.xticks(np.arange(0, 50, 1))
plt.yticks(np.arange(0, 1, 0.1))
    #plt.xlim(50, 100)
#plt.show()
#plt.savefig('./Distribution_Figure/task_Avg_acc.png')

print(task_acc)

import openpyxl
wb = openpyxl.load_workbook('Acc_result.xlsx')
sheet = wb['工作表1']
c = sheet.cell(row = 1, column = 1)
j = 1
while c.value is not None:
    j += 1
    c = sheet.cell(row=1, column = j)
#j += 1
for i in range(len(task_acc)):
    sheet.cell(row = i+2, column = j, value = task_acc[i])

wb.create_sheet('history_acc')
sh = wb['history_acc']

for i in range(51):
    sh.cell(row = 1, column = i+1, value = 'task' + str(i+1))

for i in range(51):
    for k in range(51):
        sh.cell(row = i+2, column = k+1, value = history_task_acc[i][k])
    
wb.save('Acc_result.xlsx')

